#-*- coding: utf-8 -*-
import openpyxl
import codecs
from chinese_stroke_sorting import sort_by_stroke

# 解析xlsx并创建行列表
# xlsx = ["109普高2017届高三毕业生名册.xlsx","109普高2018届高三毕业生名册.xlsx","109普高2019届高三毕业生名册.xlsx","109普高2020届高三毕业生名册.xlsx"]
with open("class_student_name_mg.txt", "w", encoding='utf-8-sig') as txtfile:
    txtfile.close()
# for xlsx_file_name in xlsx:
#     print(xlsx_file_name)
workbook=openpyxl.load_workbook("109美高2017届-2020届高三毕业生名册（教务）.xlsx")
shenames=workbook.sheetnames
# for shename in shenames:
for shename in shenames:
    worksheet=workbook[shename]
    rows=worksheet.max_row
    rowlist = []
    for row in worksheet.rows:
        celllist = []
        for cell in row:
            celllist.append(str(cell.value))
        rowlist.append(celllist)
    rowlist.pop(0)
    rowlist.pop(0)
    # rowlist.pop(0)
    # rowlist.pop(0)
    print(rowlist)
    # exit()
    # 创建基础字典
    clss_list = []
    for box in rowlist:
        try :
            clss_list.append(int(box[1]))
        except:
            print(str(box) + " ERROR ")
            pass
        # if int(box[1]) is False:
        #     pass
        # else:
        #     clss_list.append(int(box[1]))
    clss_set = list(set(clss_list))
    clss_dict = {}
    for num in clss_set:

        clss_dict[num] = []
    # 向字典内添加姓名
    for student in rowlist:
        try:
            clss = int(student[1])
            name = student[2]
            clss_dict[clss].append(name)
        except:
            print(str(student) + " ERROR ")
            pass
        # if int(student[1]) == False:
        #     pass
        # else:
        #     clss = int(student[1])
        #     name = student[3]
        #     clss_dict[clss].append(name)
    # print(clss_dict)

    # 姓名列表按照比划排序
    for i in clss_set:
        clss_dict[i] = sort_by_stroke(clss_dict[i])
        utils_list = []
        for name in clss_dict[i]:
            if len(name) == 2:
                name = name[0] + "  " + name[1]
            utils_list.append(name)
        clss_dict[i] = utils_list
    # print(clss_dict)

    # 写入txt文件
    index2 = True
    with open("class_student_name_mg.txt", "a", encoding='utf-8-sig') as txtfile:
        # with open("class_student_name.txt", "a") as txtfile:
        #     txtfile.write(xlsx_file_name.split("届")[0] + "届初三" + str(i) + "班  " + str(len(clss_dict[i])) + " 人  班主任  " + "\n")
        if index2 == True:
            # txtfile.write(xlsx_file_name.split("届")[0].split("高")[1] + "年届高中毕业" + "\n")
            # print(shenames[i])
            # exit()
            txtfile.write(shename + "年届美术高中毕业" + "\n")
            index2 = False
    for i in clss_set:
        with open("class_student_name_mg.txt","a", encoding='utf-8-sig') as txtfile:
        # with open("class_student_name.txt", "a") as txtfile:
        #     txtfile.write(xlsx_file_name.split("届")[0] + "届初三" + str(i) + "班  " + str(len(clss_dict[i])) + " 人  班主任  " + "\n")
        #     if index2 == True:
        #         txtfile.write(xlsx_file_name.split("届")[0] + "年届" + "\n")
        #         index2 = False
            txtfile.write("高三" + str(i) + "班  " + str(len(clss_dict[i])) + " 人  班主任  " + "\n")
            index = 0
            for name in clss_dict[i]:
                txtfile.write(str(name + "  "))
                index += 1
                if index == 12:
                    txtfile.write("\n")
                    index = 0
                if name == clss_dict[i][-1]:
                    txtfile.write("\n")
                    txtfile.write("\n")
            txtfile.close()
    index2 = True